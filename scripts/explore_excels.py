"""
Discovery script for the two source Excels.
Goal: understand structure before designing schema/migration.
Run: python3 scripts/explore_excels.py
"""

from pathlib import Path
import openpyxl
import pandas as pd
import sys

DESKTOP = Path("/Users/felipevelloso/Desktop")
AUTH_PATH = DESKTOP / "Autorização_de_COMPRAS_SEMANA.08.09.25_ATE_12.09.25_ATUALIZ05.09.25_17.45(1) (3) (3).xlsx"
ERP_PATH = DESKTOP / "Compras_Out_2025.xlsx"


def section(title: str) -> None:
    print("\n" + "=" * 70)
    print(title)
    print("=" * 70)


def explore_auth_file() -> None:
    section(f"AUTORIZAÇÃO — {AUTH_PATH.name}")
    if not AUTH_PATH.exists():
        print(f"NOT FOUND: {AUTH_PATH}")
        return
    print(f"Size: {AUTH_PATH.stat().st_size / 1024 / 1024:.1f} MB")

    wb = openpyxl.load_workbook(AUTH_PATH, read_only=True, data_only=True)
    sheets = wb.sheetnames
    print(f"Total sheets: {len(sheets)}")
    print(f"All sheet names:")
    for i, s in enumerate(sheets):
        print(f"  [{i:02d}] {s!r}")

    # BASE sheet — try a few likely names
    base_candidates = [s for s in sheets if s.strip().upper() in ("BASE", "BASES", "CADASTRO")]
    print(f"\nBASE candidates: {base_candidates}")

    if base_candidates:
        base_name = base_candidates[0]
        ws = wb[base_name]
        print(f"\nBASE sheet ('{base_name}') dims: {ws.max_row} rows × {ws.max_column} cols")
        # Read first 20 rows to inspect headers
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            rows.append(row)
            if i >= 20:
                break
        print("First 20 rows of BASE:")
        for r in rows:
            print(" ", r)

    # Sample a week sheet — anything that looks like a date
    week_candidates = [s for s in sheets if any(c.isdigit() for c in s) and "." in s]
    print(f"\nWeek sheet candidates (first 10): {week_candidates[:10]}")
    print(f"Total week-like sheets: {len(week_candidates)}")

    if week_candidates:
        sample = week_candidates[0]
        ws = wb[sample]
        print(f"\nSample week sheet ('{sample}') dims: {ws.max_row} × {ws.max_column}")
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            rows.append(row)
            if i >= 15:
                break
        print(f"First 15 rows of '{sample}':")
        for r in rows:
            print(" ", r)

    wb.close()


def explore_erp_file() -> None:
    section(f"ERP — {ERP_PATH.name}")
    if not ERP_PATH.exists():
        print(f"NOT FOUND: {ERP_PATH}")
        return
    print(f"Size: {ERP_PATH.stat().st_size / 1024 / 1024:.1f} MB")

    wb = openpyxl.load_workbook(ERP_PATH, read_only=True, data_only=True)
    print(f"Sheets: {wb.sheetnames}")

    target = None
    for s in wb.sheetnames:
        if "Filtrado" in s and "Alfa" in s:
            target = s
            break
    print(f"Target sheet: {target!r}")

    if target:
        ws = wb[target]
        print(f"Dims: {ws.max_row} × {ws.max_column}")
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            rows.append(row)
            if i >= 25:
                break
        print("First 25 rows:")
        for r in rows:
            print(" ", r)

    wb.close()


if __name__ == "__main__":
    explore_auth_file()
    explore_erp_file()
