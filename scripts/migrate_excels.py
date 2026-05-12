"""
Script de migração dos dados antigos do Excel pro Supabase.

Fases:
  1. Lê a última semana ('04-09 ) (3)') do Excel de autorização e importa como itens.
  2. Faz fuzzy match dos nomes dos itens contra códigos Queóps do Excel do ERP.
  3. (Opcional) Importa as últimas 4 abas semanais como solicitações históricas.
     Requer --comprador-id (UUID de um profile ativo).

Uso:
  python3 scripts/migrate_excels.py                       # fases 1+2
  python3 scripts/migrate_excels.py --comprador-id UUID   # fases 1+2+3
  python3 scripts/migrate_excels.py --reset               # apaga itens + solicitações antes (cuidado)

Saída:
  scripts/_migration_report.txt
  scripts/_migration_pending_codes.csv   (itens sem código Queóps após fuzzy match)
"""

from __future__ import annotations

from pathlib import Path
from collections import Counter
from typing import Optional
import argparse
import csv
import json
import re
import sys
import urllib.error
import urllib.request

import openpyxl
from rapidfuzz import fuzz, process

# =============================================================
# Config
# =============================================================
DESKTOP = Path("/Users/felipevelloso/Desktop")
AUTH_PATH = DESKTOP / "Autorização_de_COMPRAS_SEMANA.08.09.25_ATE_12.09.25_ATUALIZ05.09.25_17.45(1) (3) (3).xlsx"
ERP_PATH = DESKTOP / "Compras_Out_2025.xlsx"

ROOT = Path(__file__).parent.parent
ENV_PATH = ROOT / ".env.local"
REPORT_PATH = ROOT / "scripts" / "_migration_report.txt"
PENDING_CSV_PATH = ROOT / "scripts" / "_migration_pending_codes.csv"

LAST_WEEK_SHEET = "04-09   ) (3)"
LAST_4_WEEKS = ["14-08   (2)", "21-08   )", "28-08   ) (2)", "04-09   ) (3)"]

# Layout das colunas em uma aba semanal (1-indexed):
# Headers na linha 7: STATUS | Data Aprov | Data Compra | Semana Rec | Data Rec |
#   Produto Rec | Classificação | ITEM | Volume Estoque | Volume Solicitado |
#   Unidade | Preço | Forma Pagto | Prazo | Vencimento | Fornecedor | VALOR | Obs
COL = {
    "num": 2, "status": 3, "data_aprov": 4, "data_compra": 5, "semana_rec": 6,
    "data_rec": 7, "produto_rec": 8, "classif": 9, "item": 10,
    "volume_estoque": 11, "volume_solicitado": 12,
    "unidade": 13, "preco": 14, "pagto": 15, "prazo": 16, "venc": 17,
    "fornecedor": 18, "valor": 19, "obs": 20,
}

# Mapeamento de normalização
UNIDADE_MAP = {
    "KG": "KG", "kg": "KG",
    "LITRO": "LITRO", "LT": "LITRO", "L": "LITRO",
    "UNIDADE": "UNIDADE", "UN": "UNIDADE", "UND": "UNIDADE", "UNI": "UNIDADE", "PEÇA": "UNIDADE", "PECA": "UNIDADE",
    "CAIXA": "CAIXA", "CX": "CAIXA",
    "FARDO": "FARDO", "FD": "FARDO",
    "PACOTE": "PACOTE", "PCT": "PACOTE", "BAG": "PACOTE",
    "ROLO": "ROLO",
    "BANDEJA": "BANDEJA", "BDJ": "BANDEJA",
    "BISNAGA": "BISNAGA",
    "BOMBONA": "BOMBONA",
    "GARRAFA": "GARRAFA", "GFA": "GARRAFA",
    "MILHEIRO": "MILHEIRO",
    "BALDE": "BALDE",
}
CLASSIF_MAP = {
    "PAES": "PÃES", "PÃES": "PÃES",
    "CODIMENTOS": "CONDIMENTOS",
    "GORDURA VEGETAL": "GORDURA VEGETAL",
}
PAGTO_MAP = {
    "BOLETO": "BOLETO", "boleto": "BOLETO",
    "PIX": "PIX", "pix": "PIX",
    "DINHEIRO": "DINHEIRO",
    "VISA": "CARTÃO VISA", "CARTAO VISA": "CARTÃO VISA", "CARTÃO VISA": "CARTÃO VISA",
    "ELO": "CARTÃO ELO", "CARTÃO ELO": "CARTÃO ELO", "CARTAO ELO": "CARTÃO ELO",
    "NUBANK": "CARTÃO NUBANK", "CARTÃO NUBANK": "CARTÃO NUBANK", "CARTAO NUBANK": "CARTÃO NUBANK",
    "BRADESCO": "CARTÃO BRADESCO", "CARTÃO BRADESCO": "CARTÃO BRADESCO", "CARTAO BRADESCO": "CARTÃO BRADESCO",
}


# =============================================================
# Env / Supabase
# =============================================================
def load_env() -> dict:
    env = {}
    for line in ENV_PATH.read_text().splitlines():
        if "=" in line and not line.startswith("#"):
            k, v = line.split("=", 1)
            env[k.strip()] = v.strip().strip('"').strip("'")
    return env


class Supabase:
    """PostgREST client usando service_role key (bypassa RLS)."""

    def __init__(self, url: str, secret: str):
        self.url = url.rstrip("/")
        self.secret = secret

    def _request(self, method: str, path: str, body=None, params: Optional[dict] = None, prefer: Optional[str] = None):
        url = f"{self.url}{path}"
        if params:
            from urllib.parse import urlencode
            url += "?" + urlencode(params)
        headers = {
            "apikey": self.secret,
            "Authorization": f"Bearer {self.secret}",
            "Content-Type": "application/json",
            "User-Agent": "panas-compras-cli/0.1",
        }
        if prefer:
            headers["Prefer"] = prefer
        data = json.dumps(body).encode() if body is not None else None
        req = urllib.request.Request(url, data=data, headers=headers, method=method)
        try:
            with urllib.request.urlopen(req, timeout=120) as resp:
                text = resp.read().decode()
                return json.loads(text) if text else None
        except urllib.error.HTTPError as e:
            err_body = e.read().decode()
            raise RuntimeError(f"HTTP {e.code} {method} {path}: {err_body}") from e

    def select(self, table: str, params: Optional[dict] = None) -> list:
        return self._request("GET", f"/rest/v1/{table}", params=params)

    def insert(self, table: str, rows: list, on_conflict: Optional[str] = None) -> list:
        prefer = "return=representation"
        params = {}
        if on_conflict:
            params["on_conflict"] = on_conflict
            prefer += ",resolution=merge-duplicates"
        return self._request("POST", f"/rest/v1/{table}", body=rows, params=params, prefer=prefer)

    def update(self, table: str, where_eq: dict, patch: dict) -> list:
        params = {}
        for k, v in where_eq.items():
            params[k] = f"eq.{v}"
        return self._request("PATCH", f"/rest/v1/{table}", body=patch, params=params, prefer="return=representation")

    def rpc_sql(self, sql: str) -> None:
        """Roda SQL via Management API (precisa SUPABASE_ACCESS_TOKEN sbp_)."""
        import os as _os
        token = _os.environ.get("SUPABASE_ACCESS_TOKEN") or load_env().get("SUPABASE_ACCESS_TOKEN", "")
        if not token:
            raise RuntimeError("SUPABASE_ACCESS_TOKEN ausente — necessário pra rodar SQL admin.")
        project_ref = self.url.replace("https://", "").split(".")[0]
        url = f"https://api.supabase.com/v1/projects/{project_ref}/database/query"
        headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json",
            "User-Agent": "panas-compras-cli/0.1",
        }
        data = json.dumps({"query": sql}).encode()
        req = urllib.request.Request(url, data=data, headers=headers, method="POST")
        try:
            urllib.request.urlopen(req, timeout=120).read()
        except urllib.error.HTTPError as e:
            raise RuntimeError(f"SQL admin falhou: {e.read().decode()}") from e


# =============================================================
# Excel helpers
# =============================================================
def normalize_str(v) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip()


def looks_like_garbage(s: str) -> bool:
    """Rejeita valores que claramente não são uma label/nome — números, datas, vazios."""
    if not s:
        return True
    if re.fullmatch(r"[\d.,]+", s):
        return True
    if re.fullmatch(r"\d{1,2}/\d{1,2}(/\d{2,4})?", s):
        return True
    if re.search(r"\d{2}/\d{2}", s) and "-" in s:
        return True
    if len(s) > 100:
        return True
    return False


def normalize_unidade(s: str) -> str:
    key = s.strip().upper().rstrip("S")
    return UNIDADE_MAP.get(s.strip(), UNIDADE_MAP.get(key, s.strip().upper()))


def normalize_classif(s: str) -> str:
    s2 = s.strip().upper()
    return CLASSIF_MAP.get(s.strip(), CLASSIF_MAP.get(s2, s2))


def normalize_pagto(s: str) -> str:
    s2 = s.strip().upper()
    return PAGTO_MAP.get(s.strip(), PAGTO_MAP.get(s2, s2))


def is_data_row(row) -> bool:
    """Aceita linha se tem ITEM preenchido (não header, não vazia)."""
    if not row or len(row) < COL["item"]:
        return False
    item = row[COL["item"] - 1]
    if item is None:
        return False
    if isinstance(item, str) and item.strip().upper() == "ITEM":
        return False
    return True


def parse_queops_code(val) -> Optional[str]:
    if val is None:
        return None
    s = str(val).strip()
    if re.fullmatch(r"\d{5,7}", s):
        return s
    return None


def to_float(val) -> Optional[float]:
    if val is None or val == "":
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).replace(",", ".").strip()
    try:
        return float(s)
    except ValueError:
        return None


def to_iso_date(val) -> Optional[str]:
    """Aceita datetime/date ou string dd/mm/yyyy."""
    if val is None or val == "":
        return None
    from datetime import datetime, date as DateT
    if isinstance(val, (datetime, DateT)):
        return val.strftime("%Y-%m-%d") if hasattr(val, "strftime") else None
    s = str(val).strip()
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{2,4})$", s)
    if m:
        d, mo, y = m.groups()
        if len(y) == 2:
            y = "20" + y
        return f"{y}-{int(mo):02d}-{int(d):02d}"
    return None


def parse_week_dates(name: str, default_year: int = 2025) -> tuple[str, str]:
    """'04-09 ) (3)' → ('2025-09-04', '2025-09-08'). Assume semana de 5 dias úteis."""
    from datetime import date, timedelta
    m = re.search(r"(\d{1,2})-(\d{1,2})", name)
    if not m:
        return f"{default_year}-01-01", f"{default_year}-01-05"
    d, mo = int(m.group(1)), int(m.group(2))
    start = date(default_year, mo, d)
    end = start + timedelta(days=4)
    return start.isoformat(), end.isoformat()


# =============================================================
# Fases
# =============================================================
def phase1_import_items(sb: Supabase, sheet_name: str) -> dict[str, str]:
    """Retorna mapping nome -> item_id."""
    print(f"\n>>> FASE 1: importando itens da aba '{sheet_name}'")
    wb = openpyxl.load_workbook(AUTH_PATH, read_only=True, data_only=True)
    ws = wb[sheet_name]

    # Carrega lookups existentes
    classifs = {c["nome"]: c["id"] for c in sb.select("classificacoes", {"select": "id,nome"})}
    unidades = {u["nome"]: u["id"] for u in sb.select("unidades_medida", {"select": "id,nome"})}
    fornecedores = {f["nome"]: f["id"] for f in sb.select("fornecedores", {"select": "id,nome"})}
    formas_pagto = {p["nome"]: p["id"] for p in sb.select("formas_pagamento", {"select": "id,nome"})}

    items_to_insert = []
    seen_names = set()
    skipped_no_name = 0
    new_classifs: set[str] = set()
    new_fornecedores: set[str] = set()
    new_unidades: set[str] = set()
    new_pagtos: set[str] = set()

    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i < 7 or not is_data_row(row):
            continue
        nome = normalize_str(row[COL["item"] - 1])
        if not nome:
            skipped_no_name += 1
            continue
        if nome.upper() in seen_names:
            continue
        seen_names.add(nome.upper())

        classif_raw = normalize_str(row[COL["classif"] - 1])
        unidade_raw = normalize_str(row[COL["unidade"] - 1])
        forn_raw = normalize_str(row[COL["fornecedor"] - 1])
        pagto_raw = normalize_str(row[COL["pagto"] - 1])

        classif = normalize_classif(classif_raw) if classif_raw and not looks_like_garbage(classif_raw) else None
        unidade = normalize_unidade(unidade_raw) if unidade_raw and not looks_like_garbage(unidade_raw) else None
        pagto = normalize_pagto(pagto_raw) if pagto_raw and not looks_like_garbage(pagto_raw) else None
        fornecedor = forn_raw if forn_raw and not looks_like_garbage(forn_raw) else None

        if classif and classif not in classifs:
            new_classifs.add(classif)
        if unidade and unidade not in unidades:
            new_unidades.add(unidade)
        if fornecedor and fornecedor not in fornecedores:
            new_fornecedores.add(fornecedor)
        if pagto and pagto not in formas_pagto:
            new_pagtos.add(pagto)

        items_to_insert.append({
            "_nome": nome,
            "_classif": classif,
            "_unidade": unidade,
            "_fornecedor": fornecedor,
            "_pagto": pagto,
            "_preco": to_float(row[COL["preco"] - 1]),
            "_prazo": normalize_str(row[COL["prazo"] - 1]) or None,
        })

    wb.close()

    # Insere lookups novos
    if new_classifs:
        print(f"  → criando {len(new_classifs)} classificações novas: {sorted(new_classifs)}")
        rows = sb.insert("classificacoes", [{"nome": n} for n in sorted(new_classifs)], on_conflict="nome")
        for r in rows:
            classifs[r["nome"]] = r["id"]
    if new_unidades:
        print(f"  → criando {len(new_unidades)} unidades novas: {sorted(new_unidades)}")
        rows = sb.insert("unidades_medida", [{"nome": n} for n in sorted(new_unidades)], on_conflict="nome")
        for r in rows:
            unidades[r["nome"]] = r["id"]
    if new_fornecedores:
        print(f"  → criando {len(new_fornecedores)} fornecedores novos: {sorted(new_fornecedores)}")
        rows = sb.insert("fornecedores", [{"nome": n} for n in sorted(new_fornecedores)], on_conflict="nome")
        for r in rows:
            fornecedores[r["nome"]] = r["id"]
    if new_pagtos:
        print(f"  → criando {len(new_pagtos)} formas de pagto novas: {sorted(new_pagtos)}")
        rows = sb.insert("formas_pagamento", [{"nome": n} for n in sorted(new_pagtos)], on_conflict="nome")
        for r in rows:
            formas_pagto[r["nome"]] = r["id"]

    # Pre-check itens já existentes pra não duplicar
    existing = sb.select("itens", {"select": "id,nome"})
    name_to_id = {r["nome"].upper(): r["id"] for r in existing}

    items_payload = []
    skipped_existing = 0
    for it in items_to_insert:
        if it["_nome"].upper() in name_to_id:
            skipped_existing += 1
            continue
        items_payload.append({
            "nome": it["_nome"],
            "classificacao_id": classifs.get(it["_classif"]) if it["_classif"] else None,
            "unidade_id": unidades.get(it["_unidade"]) if it["_unidade"] else None,
            "fornecedor_padrao_id": fornecedores.get(it["_fornecedor"]) if it["_fornecedor"] else None,
            "forma_pagto_padrao_id": formas_pagto.get(it["_pagto"]) if it["_pagto"] else None,
            "preco_referencia": it["_preco"],
            "prazo_padrao": it["_prazo"],
            "ativo": True,
        })

    print(f"  → {skipped_existing} já existentes (pulados); inserindo {len(items_payload)} novos...")
    if items_payload:
        # Insere em lotes pra evitar timeouts
        BATCH = 100
        for i in range(0, len(items_payload), BATCH):
            inserted = sb.insert("itens", items_payload[i:i+BATCH])
            if inserted:
                for r in inserted:
                    name_to_id[r["nome"].upper()] = r["id"]

    # Reconstruir name_to_id com case-sensitive nomes (necessário pro phase 3 lookup)
    fresh = sb.select("itens", {"select": "id,nome"})
    nome_to_id = {r["nome"]: r["id"] for r in fresh}
    print(f"  → total no banco: {len(nome_to_id)} itens (skipped sem nome: {skipped_no_name})")
    return nome_to_id


def phase2_match_queops(sb: Supabase, name_to_id: dict[str, str], threshold: int = 85) -> dict:
    print(f"\n>>> FASE 2: fuzzy match de códigos Queóps (threshold {threshold})")
    # Limpa códigos existentes pra re-aplicar (idempotente)
    sb.rpc_sql("UPDATE public.itens SET codigo_queops = NULL")
    wb = openpyxl.load_workbook(ERP_PATH, read_only=True, data_only=True)
    ws = wb["Filtrado por ordem Alfabetica"]

    code_to_desc: dict[str, str] = {}
    for row in ws.iter_rows(values_only=True):
        if not row or len(row) < 4:
            continue
        code = parse_queops_code(row[1])
        if not code:
            continue
        desc = row[3]
        if not isinstance(desc, str):
            continue
        desc_clean = desc.strip()
        # Pula linhas que são "Fornecedor" e não item (têm CNPJ / parens muito longos)
        if "CNPJ" in desc_clean.upper() or len(desc_clean) > 80:
            continue
        if code not in code_to_desc:
            code_to_desc[code] = desc_clean
    wb.close()

    desc_list = list(code_to_desc.values())
    code_list = list(code_to_desc.keys())
    print(f"  → {len(code_to_desc)} códigos Queóps disponíveis")

    desc_upper = [d.upper() for d in desc_list]
    raw_matches = []  # candidatos (score >= threshold)
    pending_pre = []

    for nome, item_id in name_to_id.items():
        # Tenta token_sort + partial_ratio, fica com o maior
        m1 = process.extractOne(nome.upper(), desc_upper, scorer=fuzz.token_sort_ratio)
        m2 = process.extractOne(nome.upper(), desc_upper, scorer=fuzz.partial_ratio)
        best = max([m for m in (m1, m2) if m], key=lambda m: m[1])
        if best and best[1] >= threshold:
            idx = best[2]
            raw_matches.append({
                "item_id": item_id,
                "nome": nome,
                "codigo_queops": code_list[idx],
                "matched_desc": desc_list[idx],
                "score": best[1],
            })
        else:
            pending_pre.append({
                "nome": nome,
                "item_id": item_id,
                "best_guess": best[0] if best else None,
                "score": best[1] if best else 0,
            })

    # Dedupe por codigo_queops: o melhor score ganha o código, perdedores vão pra pending
    raw_matches.sort(key=lambda r: -r["score"])
    taken_codes: set[str] = set()
    matched = []
    losers = []
    for r in raw_matches:
        if r["codigo_queops"] in taken_codes:
            losers.append({"nome": r["nome"], "item_id": r["item_id"], "best_guess": r["matched_desc"], "score": r["score"]})
            continue
        taken_codes.add(r["codigo_queops"])
        matched.append(r)

    pending = pending_pre + losers
    print(f"  → {len(matched)} matches únicos; {len(pending)} pendentes (incluindo {len(losers)} conflitos)")

    if matched:
        for m in matched:
            sb.update("itens", {"id": m["item_id"]}, {"codigo_queops": m["codigo_queops"]})
        print(f"  → códigos atualizados em itens")

    return {"matched": matched, "pending": pending}


def phase3_import_historical(sb: Supabase, comprador_id: str, weeks: list[str], item_name_to_id: dict[str, str]):
    print(f"\n>>> FASE 3: importando {len(weeks)} semanas históricas (comprador: {comprador_id})")
    wb = openpyxl.load_workbook(AUTH_PATH, read_only=True, data_only=True)

    fornecedores = {f["nome"]: f["id"] for f in sb.select("fornecedores", {"select": "id,nome"})}
    formas_pagto = {p["nome"]: p["id"] for p in sb.select("formas_pagamento", {"select": "id,nome"})}

    summary = []
    for week_name in weeks:
        ws = wb[week_name]
        data_inicio, data_fim = parse_week_dates(week_name)
        # Cria solicitação
        result = sb.insert("solicitacoes_semanais", [{
            "data_inicio": data_inicio,
            "data_fim": data_fim,
            "comprador_id": comprador_id,
            "observacoes": f"Importado da aba '{week_name}'",
            "finalizada": True,
            "finalizada_em": f"{data_fim}T00:00:00Z",
        }])
        solic_id = result[0]["id"]

        linhas = []
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if i < 7 or not is_data_row(row):
                continue
            nome = normalize_str(row[COL["item"] - 1])
            if not nome:
                continue
            item_id = item_name_to_id.get(nome)
            if not item_id:
                # Item da semana antiga que não está no cadastro novo — pula
                continue

            forn_name = normalize_str(row[COL["fornecedor"] - 1])
            pagto_name = normalize_pagto(normalize_str(row[COL["pagto"] - 1]))

            status_raw = normalize_str(row[COL["status"] - 1]).upper()
            status_map = {
                "PARA APROVAR": "Para Aprovar",
                "AG APROVAÇÃO": "Para Aprovar",
                "APROVADO": "Aprovada",
                "APROVADO & RECEBIDO": "Aprovada & Recebida",
                "RECUSADO": "Recusada",
                "ALTERADO & APROVADO": "Volumes ou Preço Alterados",
            }
            status = status_map.get(status_raw, "Para Aprovar")

            linhas.append({
                "solicitacao_id": solic_id,
                "item_id": item_id,
                "volume_estoque": to_float(row[COL["volume_estoque"] - 1]),
                "volume_solicitado": to_float(row[COL["volume_solicitado"] - 1]) or 0,
                "preco": to_float(row[COL["preco"] - 1]) or 0,
                "fornecedor_id": fornecedores.get(forn_name),
                "forma_pagto_id": formas_pagto.get(pagto_name),
                "prazo": normalize_str(row[COL["prazo"] - 1]) or None,
                "data_compra": to_iso_date(row[COL["data_compra"] - 1]),
                "data_recebimento": to_iso_date(row[COL["data_rec"] - 1]),
                "status": status,
            })

        if linhas:
            # Insere em lotes de 100
            BATCH = 100
            for i in range(0, len(linhas), BATCH):
                sb.insert("solicitacao_linhas", linhas[i:i+BATCH])

        summary.append({"week": week_name, "solic_id": solic_id, "linhas": len(linhas)})
        print(f"  → '{week_name}': {len(linhas)} linhas importadas")

    wb.close()
    return summary


# =============================================================
# Main
# =============================================================
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--comprador-id", help="UUID de um profile ativo (pra fase 3)")
    parser.add_argument("--reset", action="store_true", help="Apaga itens/solicitações antes")
    parser.add_argument("--threshold", type=int, default=85)
    args = parser.parse_args()

    env = load_env()
    url = env.get("NEXT_PUBLIC_SUPABASE_URL")
    secret = env.get("SUPABASE_SECRET_KEY")
    if not url or not secret:
        sys.exit("ERRO: NEXT_PUBLIC_SUPABASE_URL / SUPABASE_SECRET_KEY ausentes no .env.local.")
    sb = Supabase(url, secret)

    if args.reset:
        print("Limpando dados antes de migrar (apaga itens e solicitações)...")
        sb.rpc_sql(
            "TRUNCATE public.audit_log, public.solicitacao_linhas, public.solicitacoes_semanais RESTART IDENTITY CASCADE; "
            "DELETE FROM public.itens;"
        )

    name_to_id = phase1_import_items(sb, LAST_WEEK_SHEET)
    match_results = phase2_match_queops(sb, name_to_id, threshold=args.threshold)

    historical_summary = None
    if args.comprador_id:
        historical_summary = phase3_import_historical(sb, args.comprador_id, LAST_4_WEEKS, name_to_id)
    else:
        print("\n(Pulei FASE 3 — sem --comprador-id)")

    # Reports
    PENDING_CSV_PATH.parent.mkdir(exist_ok=True)
    with PENDING_CSV_PATH.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["nome_item", "melhor_palpite_erp", "score"])
        for p in match_results["pending"]:
            writer.writerow([p["nome"], p["best_guess"] or "", p["score"]])

    with REPORT_PATH.open("w", encoding="utf-8") as f:
        f.write("RELATÓRIO DE MIGRAÇÃO\n")
        f.write("=" * 50 + "\n\n")
        f.write(f"Itens importados: {len(name_to_id)}\n")
        f.write(f"Códigos Queóps casados (fuzzy): {len(match_results['matched'])}\n")
        f.write(f"Itens sem código (pendentes): {len(match_results['pending'])}\n")
        if historical_summary:
            f.write("\nSemanas históricas:\n")
            for h in historical_summary:
                f.write(f"  - {h['week']}: {h['linhas']} linhas (id {h['solic_id']})\n")
        f.write(f"\nCSV de pendentes: {PENDING_CSV_PATH.name}\n")

    print(f"\nRelatório: {REPORT_PATH}")
    print(f"Pendentes (CSV): {PENDING_CSV_PATH}")
    print(f"\n  Itens: {len(name_to_id)}")
    print(f"  Casados: {len(match_results['matched'])}")
    print(f"  Pendentes: {len(match_results['pending'])}")


if __name__ == "__main__":
    main()
