"""
Deeper analysis of both Excels — counts, distinct values, duplicates.
Output: prints summary + writes JSON to scripts/_analysis.json
"""

from __future__ import annotations

from pathlib import Path
from collections import Counter, defaultdict
from typing import Optional
import openpyxl
import json
import re

DESKTOP = Path("/Users/felipevelloso/Desktop")
AUTH_PATH = DESKTOP / "Autorização_de_COMPRAS_SEMANA.08.09.25_ATE_12.09.25_ATUALIZ05.09.25_17.45(1) (3) (3).xlsx"
ERP_PATH = DESKTOP / "Compras_Out_2025.xlsx"
OUT_JSON = Path(__file__).parent / "_analysis.json"


# ---------- BASE columns (from explore output) ----------
# Row 6 (1-indexed) is the header row in BASE; data starts row 7
# Cols (1-indexed):
#   B(2)=#, C(3)=STATUS, D(4)=Data Aprov, E(5)=Data Compra, F(6)=Semana Rec,
#   G(7)=Data Rec, H(8)=Produto Rec, I(9)=Classificação, J(10)=ITEM,
#   K(11)=Volume, L(12)=Unidade, M(13)=Preço, N(14)=Forma Pagto, O(15)=Prazo,
#   P(16)=Vencimento, Q(17)=Fornecedor, R(18)=VALOR, S(19)=Observações
COL = {
    "num": 2, "status": 3, "data_aprov": 4, "data_compra": 5, "semana_rec": 6,
    "data_rec": 7, "produto_rec": 8, "classif": 9, "item": 10, "volume": 11,
    "unidade": 12, "preco": 13, "pagto": 14, "prazo": 15, "venc": 16,
    "fornecedor": 17, "valor": 18, "obs": 19,
}


def is_header_row(row) -> bool:
    """Detect the header row, status legend, or empty rows in BASE-like sheets."""
    if not row:
        return True
    item = row[COL["item"] - 1] if len(row) >= COL["item"] else None
    classif = row[COL["classif"] - 1] if len(row) >= COL["classif"] else None
    if item is None and classif is None:
        return True
    if isinstance(item, str) and item.strip().upper() == "ITEM":
        return True
    return False


def analyze_base(wb) -> dict:
    ws = wb["BASE"]
    items = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i < 7:
            continue
        if is_header_row(row):
            continue
        items.append({
            "classif": row[COL["classif"] - 1],
            "item": row[COL["item"] - 1],
            "unidade": row[COL["unidade"] - 1],
            "preco": row[COL["preco"] - 1],
            "pagto": row[COL["pagto"] - 1],
            "prazo": row[COL["prazo"] - 1],
            "fornecedor": row[COL["fornecedor"] - 1],
        })

    # Distinct values
    classifs = Counter(i["classif"] for i in items if i["classif"])
    unidades = Counter(str(i["unidade"]).strip() for i in items if i["unidade"])
    pagtos = Counter(str(i["pagto"]).strip() for i in items if i["pagto"])
    prazos = Counter(str(i["prazo"]).strip() for i in items if i["prazo"])
    fornecedores = Counter(str(i["fornecedor"]).strip() for i in items if i["fornecedor"])

    # Detect duplicates: case-insensitive / whitespace duplicates
    def dup_groups(counter):
        groups = defaultdict(list)
        for k, v in counter.items():
            norm = re.sub(r"\s+", " ", str(k).strip().upper())
            groups[norm].append((k, v))
        return {k: v for k, v in groups.items() if len(v) > 1}

    return {
        "total_items": len(items),
        "classifs": dict(classifs),
        "unidades": dict(unidades),
        "pagtos": dict(pagtos),
        "prazos": dict(prazos),
        "fornecedores": dict(fornecedores),
        "unidade_dups": dup_groups(unidades),
        "pagto_dups": dup_groups(pagtos),
        "fornecedor_dups": dup_groups(fornecedores),
        "sample_items": items[:5],
    }


def analyze_week_sheets(wb) -> dict:
    """Look at a few recent week sheets to confirm structure matches BASE."""
    sheets = wb.sheetnames
    skip = {"BASE", "Macro1", "Plan2", "Plan3"}
    week_sheets = [s for s in sheets if s not in skip]
    last_4 = week_sheets[-4:]

    per_sheet = {}
    for name in last_4:
        ws = wb[name]
        item_count = 0
        statuses = Counter()
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if i < 7:
                continue
            if is_header_row(row):
                continue
            item_count += 1
            status = row[COL["status"] - 1] if len(row) >= COL["status"] else None
            if status:
                statuses[str(status).strip()] += 1
        per_sheet[name] = {
            "rows_estimate": ws.max_row,
            "items_counted": item_count,
            "statuses": dict(statuses),
        }

    return {
        "total_week_sheets": len(week_sheets),
        "all_week_names": week_sheets,
        "last_4_analyzed": per_sheet,
    }


def parse_queops_code(val) -> Optional[str]:
    """Queóps codes look like 6 digits, sometimes with leading zeros, as strings."""
    if val is None:
        return None
    s = str(val).strip()
    if re.fullmatch(r"\d{5,7}", s):
        return s
    return None


def analyze_erp(wb) -> dict:
    """Extract (codigo, descricao) pairs from the ERP file."""
    ws = wb["Filtrado por ordem Alfabetica"]
    pairs = []  # (code, descricao)
    rows_seen = 0
    for row in ws.iter_rows(values_only=True):
        rows_seen += 1
        if not row or len(row) < 4:
            continue
        # Column B (1-indexed=2, 0-indexed=1) has code
        # Column D (1-indexed=4, 0-indexed=3) has description
        code = parse_queops_code(row[1])
        if not code:
            continue
        desc = row[3]
        if not isinstance(desc, str):
            continue
        pairs.append((code, desc.strip()))

    distinct = {}
    for code, desc in pairs:
        # Keep first description per code
        if code not in distinct:
            distinct[code] = desc

    return {
        "rows_in_sheet": rows_seen,
        "line_count_with_code": len(pairs),
        "distinct_codes": len(distinct),
        "sample_codes": dict(list(distinct.items())[:15]),
    }


def main():
    print("Opening authorization workbook (large, read_only=True)...")
    wb_auth = openpyxl.load_workbook(AUTH_PATH, read_only=True, data_only=True)
    base = analyze_base(wb_auth)
    weeks = analyze_week_sheets(wb_auth)
    wb_auth.close()

    print("Opening ERP workbook...")
    wb_erp = openpyxl.load_workbook(ERP_PATH, read_only=True, data_only=True)
    erp = analyze_erp(wb_erp)
    wb_erp.close()

    result = {"base": base, "weeks": weeks, "erp": erp}
    OUT_JSON.write_text(json.dumps(result, indent=2, default=str, ensure_ascii=False))
    print(f"\nWritten: {OUT_JSON}")

    # ---- Pretty summary ----
    print("\n" + "=" * 70)
    print("BASE (cadastro de itens)")
    print("=" * 70)
    print(f"Itens cadastrados: {base['total_items']}")
    print(f"Classificações distintas: {len(base['classifs'])}")
    for k, v in sorted(base["classifs"].items(), key=lambda x: -x[1]):
        print(f"  {k!r}: {v}")
    print(f"\nUnidades distintas: {len(base['unidades'])}")
    for k, v in sorted(base["unidades"].items()):
        print(f"  {k!r}: {v}")
    if base["unidade_dups"]:
        print("  → DUPLICATAS de unidade (mesmo nome em casing/whitespace diferente):")
        for norm, variants in base["unidade_dups"].items():
            print(f"    {norm}: {variants}")
    print(f"\nFormas de pagamento distintas: {len(base['pagtos'])}")
    for k, v in sorted(base["pagtos"].items()):
        print(f"  {k!r}: {v}")
    if base["pagto_dups"]:
        print("  → DUPLICATAS de pagto:")
        for norm, variants in base["pagto_dups"].items():
            print(f"    {norm}: {variants}")
    print(f"\nFornecedores distintos: {len(base['fornecedores'])}")
    for k, v in sorted(base["fornecedores"].items()):
        print(f"  {k!r}: {v}")
    if base["fornecedor_dups"]:
        print("  → DUPLICATAS de fornecedor:")
        for norm, variants in base["fornecedor_dups"].items():
            print(f"    {norm}: {variants}")
    print(f"\nPrazos distintos: {len(base['prazos'])}")
    for k, v in sorted(base["prazos"].items()):
        print(f"  {k!r}: {v}")

    print("\n" + "=" * 70)
    print("ABAS SEMANAIS")
    print("=" * 70)
    print(f"Total: {weeks['total_week_sheets']}")
    print(f"Últimas 4 analisadas em detalhe:")
    for name, info in weeks["last_4_analyzed"].items():
        print(f"\n  '{name}':")
        print(f"    linhas com item: {info['items_counted']}")
        print(f"    statuses: {info['statuses']}")

    print("\n" + "=" * 70)
    print("ERP (códigos Queóps)")
    print("=" * 70)
    print(f"Linhas-fato na aba 'Filtrado por ordem Alfabetica': {erp['line_count_with_code']}")
    print(f"Códigos Queóps distintos: {erp['distinct_codes']}")
    print(f"\nExemplos:")
    for code, desc in erp["sample_codes"].items():
        print(f"  {code} → {desc}")


if __name__ == "__main__":
    main()
